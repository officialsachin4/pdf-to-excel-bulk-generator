"""
PDF -> Excel batch converter.

Design goals (per client requirement):
- Accept a BATCH of PDFs in one upload, not just one file.
- Copy table data verbatim: every cell is written as the exact string tabula/
  pdfplumber extracted, with no .title()/.strip()/type coercion that could
  alter spacing, casing, leading zeros, symbols, etc.
- Sheet layout is CONFIGURABLE per request (different clients want different
  layouts), not hardcoded. Options:
    - "per_pdf"   : one sheet per source PDF, all that PDF's tables stacked
                    on it with a blank row + filename/table label between them.
    - "per_table" : one sheet per individual table, across all PDFs, named
                    "<pdf_stem>_T<n>" (Excel sheet names are capped at 31
                    chars, so long names are truncated safely).
    - "single"    : every table from every PDF stacked onto one sheet, in
                    upload order, each preceded by a small header row showing
                    its source file and table number.
- Two processing modes preserved from the original app:
    - "tablesOnly": tabula-based table extraction only (uses Java).
    - "allText"   : pdfplumber-based extraction of full text with embedded
                    tables, kept close to original behavior but fixed to
                    actually iterate per-uploaded-file instead of a single
                    global file handle, and to write verbatim string values.
"""

import os
import uuid
from typing import List, Tuple

import pandas as pd
import pdfplumber
import tabula
from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

# Two unrelated PyPI packages both install under the import name "tabula":
#   - tabula-py  (the one we need; provides tabula.read_pdf)
#   - tabula     (an unrelated package; has no read_pdf)
# If both ever get installed (e.g. someone runs `pip install tabula` instead
# of `pip install tabula-py`), whichever wins import resolution may be the
# wrong one, and tabula.read_pdf will be missing -> AttributeError at request
# time. Fail loudly and clearly at startup instead, so it's obvious what to
# fix rather than a 500 buried in a Flask traceback.
if not hasattr(tabula, 'read_pdf'):
    raise ImportError(
        "Wrong 'tabula' package is installed (it has no read_pdf attribute). "
        "This app needs 'tabula-py', not the unrelated 'tabula' package. Fix with:\n"
        "    pip uninstall tabula tabula-py -y\n"
        "    pip install tabula-py\n"
        f"Currently imported from: {getattr(tabula, '__file__', '<unknown>')}"
    )

app = Flask(__name__, template_folder='templates')

UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'pdf'}
VALID_LAYOUTS = {'per_pdf', 'per_table', 'single'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

HEADER_FONT = Font(name='Times New Roman', size=11, bold=True)
CELL_FONT = Font(name='Times New Roman', size=11)
LABEL_FONT = Font(name='Times New Roman', size=11, bold=True, italic=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='center')
HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')


def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def verbatim(value) -> str:
    """
    Force a value to its exact extracted string with no re-typing.
    pdfplumber/tabula sometimes hand back None or NaN for empty cells
    (pandas upcasts a mixed None/float column to NaN), and floats for
    numeric-looking strings. We normalize both None and NaN -> "" ;
    everything else is cast to str() unchanged so original spacing/
    symbols/leading zeros are preserved exactly as extracted.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            # tabula/pandas sometimes upcasts whole-number strings to float
            # (e.g. "007" the source had, becomes 7.0). We cannot recover the
            # original string once that's happened upstream, but we avoid
            # *adding* drift here by not appending ".0".
            return str(int(value))
    return str(value)


def safe_sheet_title(title: str, used_titles: set) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ] , must be unique."""
    for ch in r':\/?*[]':
        title = title.replace(ch, '-')
    title = title[:31] or "Sheet"
    base = title
    n = 1
    while title in used_titles:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used_titles.add(title)
    return title


# ---------------------------------------------------------------------------
# Mode: tablesOnly  — pdfplumber first, tabula fallback
# ---------------------------------------------------------------------------
#
# ROOT CAUSE of the "all data in column A" bug:
#   tabula-py uses the Java tabula library which detects tables by looking for
#   ruled lines/borders in the PDF. When a PDF has a *borderless* table (laid
#   out with whitespace and background fills, like the HR Roster), tabula falls
#   back to streaming mode and concatenates every cell on a row into a single
#   text blob in the first column -- exactly what was seen in the output.
#
# FIX: use pdfplumber as the primary extractor. pdfplumber uses whitespace/
#   column-alignment heuristics that work on both bordered AND borderless
#   tables. tabula is kept as a fallback for PDFs where pdfplumber finds
#   nothing (e.g. scanned PDFs with OCR'd borders).
# ---------------------------------------------------------------------------

def _extract_with_pdfplumber(pdf_path: str) -> List[pd.DataFrame]:
    """
    Extract tables using pdfplumber.

    Strategy:
    1. First try pdfplumber's native extract_tables() — works for PDFs where
       the table has either visible lines or pdfplumber can infer the structure.
    2. If that produces only 1-column results (everything merged, the classic
       borderless-table symptom), fall back to word-coordinate clustering:
       group words by their X midpoint into column buckets, then reconstruct
       rows by their Y position. This correctly handles borderless tables laid
       out with background fills and whitespace (like the HR Roster PDF).
    """
    results = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # ---- Attempt 1: native extract_tables ----
            raw_tables = page.extract_tables()
            good_tables = []
            for raw in (raw_tables or []):
                if not raw or len(raw) < 2:
                    continue
                col_count = max(len(row) for row in raw)
                # Only trust if more than 1 column was detected
                if col_count > 1:
                    headers = [verbatim(c) for c in raw[0]]
                    # De-duplicate blank/repeated headers
                    seen = {}
                    deduped = []
                    for h in headers:
                        key = h or ""
                        if key in seen:
                            seen[key] += 1
                            deduped.append(f"{key}_{seen[key]}" if key else f"Col_{len(deduped)+1}")
                        else:
                            seen[key] = 0
                            deduped.append(key or f"Col_{len(deduped)+1}")
                    rows = [[verbatim(c) for c in row] for row in raw[1:]]
                    good_tables.append(pd.DataFrame(rows, columns=deduped))

            if good_tables:
                results.extend(good_tables)
                continue

            # ---- Attempt 2: word-coordinate clustering for borderless tables ----
            # Extract every word with its bounding box coordinates.
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue

            # Group words into rows by their top-Y coordinate (±4 pt tolerance)
            Y_TOL = 4
            rows_by_y: dict = {}
            for w in words:
                y = round(w['top'])
                matched_y = None
                for existing_y in rows_by_y:
                    if abs(existing_y - y) <= Y_TOL:
                        matched_y = existing_y
                        break
                rows_by_y.setdefault(matched_y if matched_y is not None else y, []).append(w)

            sorted_ys = sorted(rows_by_y)
            if len(sorted_ys) < 2:
                continue

            # Discover column boundaries from the header row using gap analysis.
            # Gaps between consecutive header words fall into two populations:
            #   - Small gaps: words within the same column header ("FULL NAME")
            #   - Large gaps: space between adjacent columns
            # We find the split point between those two populations, then merge
            # words into column headers accordingly.
            header_words = sorted(rows_by_y[sorted_ys[0]], key=lambda w: w['x0'])
            if not header_words:
                continue

            # Compute all inter-word gaps on the header row
            gaps = []
            for j in range(1, len(header_words)):
                gap = header_words[j]['x0'] - header_words[j-1]['x1']
                gaps.append(gap)

            if gaps:
                # Column-break threshold: midpoint between the largest small
                # gap and smallest large gap. Simple but effective for uniform
                # table layouts. Sort gaps and look for the biggest jump.
                sorted_gaps = sorted(gaps)
                if len(sorted_gaps) > 1:
                    jump_idx = max(range(len(sorted_gaps)-1),
                                   key=lambda k: sorted_gaps[k+1] - sorted_gaps[k])
                    COL_GAP = (sorted_gaps[jump_idx] + sorted_gaps[jump_idx+1]) / 2
                else:
                    COL_GAP = sorted_gaps[0] + 1  # only one gap: two columns
                COL_GAP = max(COL_GAP, 5)  # never less than 5pt
            else:
                COL_GAP = 20  # single word in header: one column

            col_headers = []
            col_x_ranges = []
            for w in header_words:
                if col_x_ranges and (w['x0'] - col_x_ranges[-1][1]) < COL_GAP:
                    col_headers[-1] += ' ' + w['text']
                    col_x_ranges[-1] = (col_x_ranges[-1][0], w['x1'])
                else:
                    col_headers.append(w['text'])
                    col_x_ranges.append((w['x0'], w['x1']))

            if not col_headers:
                continue

            # Build one column-center list for assigning data words to columns.
            # Use the midpoint of each header's X range as the column center.
            col_mids = [(x0 + x1) / 2 for x0, x1 in col_x_ranges]

            def assign_col(word_x_mid: float) -> int:
                """Return index of the closest column center."""
                return min(range(len(col_mids)), key=lambda i: abs(col_mids[i] - word_x_mid))

            # Build data rows using the same gap threshold.
            # Within each row, words closer than COL_GAP to the previous word
            # are appended to the same cell; a larger gap means a new cell.
            data_rows = []
            for y in sorted_ys[1:]:
                row_words = sorted(rows_by_y[y], key=lambda w: w['x0'])
                cells = [''] * len(col_headers)
                prev_x1 = None
                current_ci = 0
                for w in row_words:
                    mid = (w['x0'] + w['x1']) / 2
                    if prev_x1 is not None and (w['x0'] - prev_x1) < COL_GAP:
                        # Same cell as previous word
                        cells[current_ci] = (cells[current_ci] + ' ' + w['text']).strip()
                    else:
                        # New cell — find the best matching column by X midpoint
                        current_ci = assign_col(mid)
                        cells[current_ci] = (cells[current_ci] + ' ' + w['text']).strip()
                    prev_x1 = w['x1']
                data_rows.append(cells)

            if data_rows:
                df = pd.DataFrame(data_rows, columns=col_headers)
                results.append(df)

    return results


def _extract_with_tabula(pdf_path: str) -> List[pd.DataFrame]:
    """
    Extract tables using tabula-py (Java-based, border-detection).
    Good fallback for scanned/OCR'd PDFs with visible ruled lines.
    Retries with cp1252/latin-1 if the PDF contains non-UTF-8 text.
    """
    encodings_to_try = ["utf-8", "cp1252", "latin-1"]
    last_error = None
    for encoding in encodings_to_try:
        try:
            dfs = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True, encoding=encoding)
            return dfs or []
        except UnicodeDecodeError as e:
            last_error = e
            continue
    raise UnicodeDecodeError(
        last_error.encoding, last_error.object, last_error.start,
        last_error.end, last_error.reason
    )


def extract_tables_from_pdf(pdf_path: str) -> List[pd.DataFrame]:
    """
    Smart two-stage extractor:
      1. Try pdfplumber (handles bordered AND borderless/whitespace tables).
      2. If pdfplumber finds nothing, fall back to tabula (better for
         scanned PDFs with OCR'd ruled lines).
    If both find nothing, return [].
    If a genuine encoding error occurs in both, raise ValueError naming the file.
    """
    try:
        dfs = _extract_with_pdfplumber(pdf_path)
        if dfs:
            return dfs
        # pdfplumber found no tables -- try tabula as fallback
        try:
            return _extract_with_tabula(pdf_path)
        except UnicodeDecodeError as e:
            raise ValueError(
                f"Could not decode text in '{os.path.basename(pdf_path)}' "
                f"(tried encodings: utf-8, cp1252, latin-1). "
                f"The PDF may use an unusual encoding or be corrupted. Original error: {e}"
            )
    except ValueError:
        raise
    except Exception as e:
        # pdfplumber itself failed (corrupted PDF, password-protected, etc.)
        raise ValueError(
            f"Could not extract tables from '{os.path.basename(pdf_path)}': {e}"
        )


def write_table_block(sheet, start_row: int, df: pd.DataFrame, label: str = None) -> int:
    """
    Write one DataFrame verbatim starting at start_row.
    If label is given, writes it as a small header line first.
    Returns the next free row after this block (with one blank-row gap).
    """
    row = start_row
    if label:
        cell = sheet.cell(row=row, column=1, value=label)
        cell.font = LABEL_FONT
        row += 1

    columns = list(df.columns)
    for col_num, col_name in enumerate(columns, start=1):
        header_text = verbatim(col_name)
        cell = sheet.cell(row=row, column=col_num, value=header_text)
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN
        cell.fill = HEADER_FILL
        current_width = sheet.column_dimensions[get_column_letter(col_num)].width or 0
        sheet.column_dimensions[get_column_letter(col_num)].width = max(current_width, len(header_text) + 2, 10)
    row += 1

    for _, data_row in df.iterrows():
        for col_num, value in enumerate(data_row, start=1):
            text = verbatim(value)
            cell = sheet.cell(row=row, column=col_num, value=text)
            cell.font = CELL_FONT
            cell.alignment = WRAP_ALIGN
            current_width = sheet.column_dimensions[get_column_letter(col_num)].width or 0
            sheet.column_dimensions[get_column_letter(col_num)].width = max(current_width, len(text) + 2, 10)
        row += 1

    return row + 1  # one blank row gap before next block


def write_tables_to_excel(per_pdf_tables: List[Tuple[str, List[pd.DataFrame]]], excel_path: str, layout: str,
                           failed_files: List[Tuple[str, str]] = None):
    """
    per_pdf_tables: list of (pdf_filename, [DataFrame, ...]) in upload order.
    layout: 'per_pdf' | 'per_table' | 'single'
    failed_files: list of (pdf_filename, error_message) for PDFs that could
        not be processed (e.g. encoding issues) -- surfaced as a dedicated
        sheet so the client can see exactly what's missing and why, instead
        of silently losing those PDFs from the batch.
    """
    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()

    if layout == 'per_pdf':
        for pdf_name, tables in per_pdf_tables:
            if not tables:
                continue
            stem = os.path.splitext(pdf_name)[0]
            sheet = wb.create_sheet(safe_sheet_title(stem, used_titles))
            row = 1
            for i, df in enumerate(tables, start=1):
                label = f"Table {i} (source: {pdf_name})" if len(tables) > 1 else f"Source: {pdf_name}"
                row = write_table_block(sheet, row, df, label=label)

    elif layout == 'per_table':
        for pdf_name, tables in per_pdf_tables:
            stem = os.path.splitext(pdf_name)[0]
            for i, df in enumerate(tables, start=1):
                sheet = wb.create_sheet(safe_sheet_title(f"{stem}_T{i}", used_titles))
                write_table_block(sheet, 1, df, label=f"Source: {pdf_name} | Table {i}")

    elif layout == 'single':
        sheet = wb.create_sheet("All Tables")
        row = 1
        for pdf_name, tables in per_pdf_tables:
            for i, df in enumerate(tables, start=1):
                row = write_table_block(sheet, row, df, label=f"Source: {pdf_name} | Table {i}")

    else:
        raise ValueError(f"Unknown layout: {layout}")

    if not wb.sheetnames:
        wb.create_sheet("No Tables Found")

    if failed_files:
        issues_sheet = wb.create_sheet("Conversion Issues")
        issues_sheet.cell(row=1, column=1, value="File").font = HEADER_FONT
        issues_sheet.cell(row=1, column=2, value="Error").font = HEADER_FONT
        for r, (fname, err) in enumerate(failed_files, start=2):
            issues_sheet.cell(row=r, column=1, value=fname).font = CELL_FONT
            issues_sheet.cell(row=r, column=2, value=err).font = CELL_FONT
        issues_sheet.column_dimensions['A'].width = 30
        issues_sheet.column_dimensions['B'].width = 80

    wb.save(excel_path)


# ---------------------------------------------------------------------------
# Mode: allText  (pdfplumber -> full text with embedded tables, verbatim)
# ---------------------------------------------------------------------------

def extract_pdf_content(pdf_path: str) -> Tuple[str, List[list]]:
    text_content = ""
    table_data = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content += page_text + "\n"
                tables = page.extract_tables()
                if tables:
                    table_data.extend(tables)
    except UnicodeDecodeError as e:
        raise ValueError(
            f"Could not decode text in '{os.path.basename(pdf_path)}'. "
            f"The PDF may use an unusual encoding or be corrupted. Original error: {e}"
        )
    return text_content, table_data


def write_text_and_tables_block(sheet, start_row: int, pdf_name: str, text_content: str, table_data: List[list]) -> int:
    row = start_row
    label_cell = sheet.cell(row=row, column=1, value=f"Source: {pdf_name}")
    label_cell.font = LABEL_FONT
    row += 1

    text_lines = text_content.split('\n')
    table_data_used = set()

    for line in text_lines:
        line_in_table = False
        matched_table_index = None
        for table_index, table in enumerate(table_data):
            if table_index in table_data_used:
                continue
            flat = " ".join(" ".join(verbatim(c) for c in r) for r in table)
            if line.strip() and line.strip() in flat:
                line_in_table = True
                matched_table_index = table_index
                break

        if not line_in_table:
            if line.strip():
                sheet.cell(row=row, column=1, value=line).font = CELL_FONT
                row += 1
            continue

        table = table_data[matched_table_index]
        table_data_used.add(matched_table_index)
        for r_offset, table_row in enumerate(table):
            for col_num, value in enumerate(table_row, start=2):  # column B onward
                cell = sheet.cell(row=row + r_offset, column=col_num, value=verbatim(value))
                cell.font = CELL_FONT
        row += len(table) + 1

    return row + 1


def create_excel_all_text(per_pdf_text: List[Tuple[str, str, List[list]]], layout: str) -> Workbook:
    """per_pdf_text: list of (pdf_filename, text_content, table_data)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_titles = set()

    if layout == 'per_pdf':
        for pdf_name, text_content, table_data in per_pdf_text:
            stem = os.path.splitext(pdf_name)[0]
            sheet = wb.create_sheet(safe_sheet_title(stem, used_titles))
            write_text_and_tables_block(sheet, 1, pdf_name, text_content, table_data)

    elif layout in ('single', 'per_table'):
        # "per_table" doesn't map cleanly onto interleaved text+table content,
        # so it falls back to one block per PDF on a shared sheet (closest
        # sane interpretation); each block is still clearly labeled.
        sheet = wb.create_sheet("All Content")
        row = 1
        for pdf_name, text_content, table_data in per_pdf_text:
            row = write_text_and_tables_block(sheet, row, pdf_name, text_content, table_data)

    else:
        raise ValueError(f"Unknown layout: {layout}")

    if not wb.sheetnames:
        wb.create_sheet("No Content Found")

    return wb


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

def delete_files(*file_paths):
    for path in file_paths:
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except PermissionError:
                pass


@app.route('/')
def index():
    return render_template('pdftoexcel.html')


@app.route('/upload', methods=['POST'])
def upload():
    files = request.files.getlist('pdfFiles')  # batch input field
    if not files or all(f.filename == '' for f in files):
        return redirect(url_for('index'))

    processing_option = request.form.get('processingOption', 'tablesOnly')
    layout = request.form.get('sheetLayout', 'per_pdf')
    if layout not in VALID_LAYOUTS:
        return f"Invalid sheetLayout '{layout}'. Choose one of {sorted(VALID_LAYOUTS)}.", 400

    saved_paths = []
    job_id = uuid.uuid4().hex[:8]

    try:
        for file in files:
            if file.filename == '':
                continue
            if not allowed_file(file.filename):
                return f"Invalid file '{file.filename}'. Only PDF files are allowed.", 400
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{job_id}_{filename}")
            file.save(filepath)
            saved_paths.append((filename, filepath))

        if not saved_paths:
            return redirect(url_for('index'))

        if processing_option == 'tablesOnly':
            per_pdf_tables = []
            failed_files = []
            for original_name, path in saved_paths:
                try:
                    tables = extract_tables_from_pdf(path)
                    per_pdf_tables.append((original_name, tables))
                except ValueError as e:
                    # One bad PDF (e.g. unusual encoding, corrupted) shouldn't
                    # abort the whole batch -- record it and keep going.
                    failed_files.append((original_name, str(e)))

            if not any(tables for _, tables in per_pdf_tables):
                if failed_files:
                    details = "; ".join(f"{name}: {err}" for name, err in failed_files)
                    return f"No tables could be extracted. Errors: {details}", 422
                return 'No tables found in the supplied PDF(s).'

            excel_path = os.path.join(OUTPUT_FOLDER, f"TablesOnly_{job_id}.xlsx")
            write_tables_to_excel(per_pdf_tables, excel_path, layout, failed_files=failed_files)
            return send_file(excel_path, as_attachment=True, download_name='TablesOnly.xlsx')

        elif processing_option == 'allText':
            per_pdf_text = []
            failed_files = []
            for original_name, path in saved_paths:
                try:
                    text_content, table_data = extract_pdf_content(path)
                    per_pdf_text.append((original_name, text_content, table_data))
                except ValueError as e:
                    failed_files.append((original_name, str(e)))

            if not per_pdf_text:
                details = "; ".join(f"{name}: {err}" for name, err in failed_files)
                return f"No content could be extracted. Errors: {details}", 422

            wb = create_excel_all_text(per_pdf_text, layout)
            if failed_files:
                issues_sheet = wb.create_sheet("Conversion Issues")
                issues_sheet.cell(row=1, column=1, value="File").font = HEADER_FONT
                issues_sheet.cell(row=1, column=2, value="Error").font = HEADER_FONT
                for r, (fname, err) in enumerate(failed_files, start=2):
                    issues_sheet.cell(row=r, column=1, value=fname).font = CELL_FONT
                    issues_sheet.cell(row=r, column=2, value=err).font = CELL_FONT
                issues_sheet.column_dimensions['A'].width = 30
                issues_sheet.column_dimensions['B'].width = 80
            excel_path = os.path.join(OUTPUT_FOLDER, f"TextPlusTables_{job_id}.xlsx")
            wb.save(excel_path)
            return send_file(excel_path, as_attachment=True, download_name='TextPlusTables.xlsx')

        else:
            return 'Invalid processing option', 400

    finally:
        delete_files(*(p for _, p in saved_paths))


if __name__ == "__main__":
    app.run(debug=True)